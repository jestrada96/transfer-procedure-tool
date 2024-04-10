globals().clear()
from openpyxl import load_workbook
from classes.valve import Valve
from classes.valve2 import Valve2
from classes.valve3 import Valve3
from classes.split import Split
from classes.nozzle import Nozzle
from classes.pump import Pump
from classes.dropleg import Dropleg
from classes.node import Node

wb = load_workbook(filename = 'Valves.xlsx')
ws = wb['Connections']
name_column=ws["B3:B200"]
class_column=ws["C3:C200"]
destinations=ws["D3:F200"]

#key = EIN, value = node object
inventory = {}

classes = {
        "2-Way-Valve": Valve2,
        "3-Way-Valve": Valve3,
        "Split Point": Split,
        "Nozzle": Nozzle,
        "Pump": Pump,
        "": Valve,
        "Dropleg": Dropleg,
        None: Valve 
}

for name_cell in name_column:
        if name_cell[0].value: 
            inventory[name_cell[0].value] = None

for key, class_row in zip(inventory.keys(), class_column):
    for class_cell in class_row:
        inventory[key] = classes[class_cell.value](key)

for node, destination_row in zip(inventory.values(), destinations):
    for destination_cell in destination_row:
        if destination_cell.value in inventory:
            if destination_cell.value:
                node.connect(inventory[destination_cell.value])

def getRoutes(a, b):
    src = inventory[a]
    dst = inventory[b]
    return src.routesTo(dst)

     
                