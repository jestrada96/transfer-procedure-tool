globals().clear()
from openpyxl import load_workbook
from classes.valve import Valve
from classes.valve2 import Valve2
from classes.valve3 import Valve3
from classes.split import Split
from classes.nozzle import Nozzle
from classes.pump import Pump
from classes.tankreturn import TankReturn
from classes.pit import Pit

wb = load_workbook(filename = 'MasterProcedureData.xlsx', data_only=True)

pits = {}

for row in wb["Pits"].iter_rows(min_row=3, values_only= True):
     pit = row[1]
     nace = row[2]
     nacePMID = row[3]
     pits[pit] = Pit(pit, nace, nacePMID)

for row in wb["Heaters"].iter_rows(min_row=3, values_only= True):
    pit_name = row[2]
    heater = row[1]
    if heater:
        pits[pit_name].heaters.append(heater)

for row in wb["TFSPS PMIDs"].iter_rows(min_row=3, values_only= True):
    pit_name = row[3]
    tfsps = row[1]
    tfsps_pmid = row[1]
    pits[pit_name].tfsps.append(tfsps)
    pits[pit_name].tfsps_pmid.append(tfsps_pmid)  

node_types = {
        "2-Way-Valve": Valve2,
        "3-Way-Valve": Valve3,
        "Split Point": Split,
        "Nozzle": Nozzle,
        "Pump": Pump,
        "": Valve,
        "Tank Return": TankReturn,
        None: Valve 
}

#key = EIN, value = node object
inventory = {}

cnx = wb['Connections']
conections_matrix=cnx["D3:F200"]

for row in cnx.iter_rows(min_row=3, values_only= True):
     node = row[1]
     node_type = row[2]
     inventory[node] = node_types[node_type](node)
     inventory[node].setPit(row[6])

for node, connections in zip(inventory.values(), conections_matrix):
    for connection in connections:
        if connection.value in inventory:
            if connection.value:
                node.connect(inventory[connection.value])



     
                